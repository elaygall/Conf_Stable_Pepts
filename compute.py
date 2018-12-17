import argparse
import re
from collections import defaultdict
import xlsxwriter
import openpyxl


def read_choices(filename):
    with open(filename, "r") as infile:
        return [x.strip() for x in infile.readlines()]


def read_sources(filename):
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.active
    col = ws['A']
    for i in range(len(col)):
        yield col[i].value


def get_stats(aseq, checkers):
    stat = [0]*len(aseq)

    for c in checkers:
        pps = [x.start() for x in re.finditer("(?={})".format(c), aseq)]
        for p in pps:
            for i in range(len(c)):
                stat[p+i] += 1
    return stat


def write_xls(outname, stat_map):
    wb = xlsxwriter.Workbook(outname)
    palette = [wb.add_format({"font_color": color})
               for color in ['black', 'blue', 'red', 'magenta', 'yellow']]
    ws = wb.add_worksheet()
    row = 0
    for sr, st in stat_map.items():
        write_row(ws, palette, row, sr, st)
        row += 1
    wb.close()


def write_row(ws, palette, row, source, stat):
    args = []
    for i, letter in enumerate(source):
        args.append(palette[stat[i]])
        args.append(letter)
    ws.write_rich_string(row, 0, *args)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='You know')
    parser.add_argument('sequence_file', type=str, help='where to get sequences')
    parser.add_argument('tetragram_file', type=str, help='what to search for')
    parser.add_argument('out_file', type=str, help='where to write results')
    parser.add_argument('--limit', type=int, default=100000,
                        help='limit tetragram number for testing purposes')

    args = parser.parse_args()

    choices = read_choices(args.tetragram_file)[:args.limit]
    # print(choices)
    # print(list(read_sources("res_xlsx.xlsx")))
    stat_map = {s: get_stats(s, choices) for s in read_sources(args.sequence_file)}
    write_xls(args.out_file, stat_map)


